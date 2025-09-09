"""
Generador de reportes Excel con estadísticas
Crea hojas de productos y resumen con gráficos
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from datetime import datetime
from pathlib import Path
from typing import List, Dict

from config.settings import ExcelSettings, REPORTS_DIR
from utils.logger import setup_logger


class ExcelReportGenerator:
    """Generador de reportes Excel"""
    
    def __init__(self):
        self.logger = setup_logger("ExcelGenerator")
        
        # Crear directorio de reportes
        REPORTS_DIR.mkdir(exist_ok=True)
    
    def generate_report(self, products: List[Dict], statistics: Dict) -> str:
        """
        Genera reporte Excel completo
        
        Args:
            products: Lista de productos
            statistics: Estadísticas calculadas
            
        Returns:
            str: Ruta del archivo Excel generado
        """
        try:
            self.logger.log_step(4, "Generación de reporte Excel", "INICIADO")
            
            # Generar nombre de archivo
            today = datetime.now().strftime('%Y-%m-%d')
            filename = f"Reporte_{today}.xlsx"
            filepath = REPORTS_DIR / filename
            
            # Crear workbook
            wb = Workbook()
            
            # Eliminar hoja por defecto
            wb.remove(wb.active)
            
            # Crear hojas
            self._create_products_sheet(wb, products)
            self._create_summary_sheet(wb, statistics)
            
            # Guardar archivo
            wb.save(filepath)
            
            file_size = filepath.stat().st_size
            
            self.logger.log_file_operation(
                "EXCEL_GENERATION",
                str(filepath),
                file_size,
                True
            )
            
            self.logger.log_step(
                4, "Generación de reporte Excel", "COMPLETADO",
                {"archivo": str(filepath), "productos": len(products)}
            )
            
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"Error generando reporte Excel: {str(e)}")
            raise
    
    def _create_products_sheet(self, workbook: Workbook, products: List[Dict]):
        """Crea la hoja de productos"""
        sheet_name = ExcelSettings.SHEET_NAMES[ExcelSettings.LANGUAGE]['products']
        ws = workbook.create_sheet(sheet_name)
        
        # Headers
        headers = ['ID', 'Título', 'Precio', 'Categoría', 'Descripción', 'Fecha Inserción']
        ws.append(headers)
        
        # Aplicar estilo a headers
        header_font = Font(bold=True, color='000000')
        header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        header_alignment = Alignment(horizontal='center')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Datos de productos
        for product in products:
            ws.append([
                product['id'],
                product['title'],
                product['price'],
                product['category'],
                product['description'][:100] + '...' if len(product['description']) > 100 else product['description'],
                product['fecha_insercion']
            ])
        
        # Ajustar anchos de columna
        column_widths = [10, 40, 15, 20, 50, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        # Formato de números para precios
        for row in range(2, len(products) + 2):
            price_cell = ws.cell(row=row, column=3)
            price_cell.number_format = '#,##0.00'
    
    def _create_summary_sheet(self, workbook: Workbook, statistics: Dict):
        """Crea la hoja de resumen con estadísticas"""
        sheet_name = ExcelSettings.SHEET_NAMES[ExcelSettings.LANGUAGE]['summary']
        ws = workbook.create_sheet(sheet_name)
        
        # Título
        ws['A1'] = 'RESUMEN DE PRODUCTOS'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
        
        # Estadísticas generales
        ws['A3'] = 'Estadísticas Generales'
        ws['A3'].font = Font(bold=True, size=14)
        
        ws['A4'] = 'Total de productos:'
        ws['B4'] = statistics.get('total_products', 0)
        
        ws['A5'] = 'Precio promedio general:'
        ws['B5'] = statistics.get('avg_price', 0)
        ws['B5'].number_format = '$#,##0.00'
        
        # Estadísticas por categoría
        ws['A7'] = 'Estadísticas por Categoría'
        ws['A7'].font = Font(bold=True, size=14)
        
        # Headers para tabla de categorías
        category_headers = ['Categoría', 'Cantidad', 'Precio Promedio', 'Precio Mínimo', 'Precio Máximo']
        for col_num, header in enumerate(category_headers, 1):
            cell = ws.cell(row=8, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        
        # Datos por categoría
        category_stats = statistics.get('category_stats', [])
        for row_num, category in enumerate(category_stats, 9):
            ws.cell(row=row_num, column=1, value=category['category'])
            ws.cell(row=row_num, column=2, value=category['count'])
            ws.cell(row=row_num, column=3, value=category['avg_price'])
            ws.cell(row=row_num, column=4, value=category['min_price'])
            ws.cell(row=row_num, column=5, value=category['max_price'])
            
            # Formato de números
            for col in [3, 4, 5]:
                ws.cell(row=row_num, column=col).number_format = '$#,##0.00'
        
        # Ajustar anchos
        column_widths = [20, 12, 18, 15, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=8, column=col_num).column_letter].width = width
        
        # Agregar gráfico si está habilitado
        if ExcelSettings.INCLUDE_CHARTS and category_stats:
            self._add_category_chart(ws, category_stats)
        
        # Información de generación
        ws[f'A{len(category_stats) + 12}'] = f'Reporte generado: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws[f'A{len(category_stats) + 12}'].font = Font(italic=True, size=10)
    
    def _add_category_chart(self, worksheet, category_stats: List[Dict]):
        """Agrega gráfico de barras por categoría"""
        try:
            # Crear gráfico de barras
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Productos por Categoría"
            chart.y_axis.title = 'Cantidad'
            chart.x_axis.title = 'Categoría'
            
            # Datos para el gráfico
            data = Reference(worksheet, min_col=2, min_row=8, max_row=8 + len(category_stats), max_col=2)
            cats = Reference(worksheet, min_col=1, min_row=9, max_row=8 + len(category_stats))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Posicionar gráfico
            chart.height = 10
            chart.width = 15
            worksheet.add_chart(chart, f"G8")
            
        except Exception as e:
            self.logger.warning(f"Error agregando gráfico: {str(e)}")


def test_excel_generator():
    """Función de prueba para el generador Excel"""
    logger = setup_logger("ExcelTest")
    
    try:
        generator = ExcelReportGenerator()
        
        # Datos de prueba
        test_products = [
            {
                'id': 1,
                'title': 'Producto Test 1',
                'price': 19.99,
                'category': 'electronics',
                'description': 'Descripción del producto test',
                'fecha_insercion': datetime.now()
            },
            {
                'id': 2,
                'title': 'Producto Test 2',
                'price': 29.99,
                'category': 'clothing',
                'description': 'Otra descripción de prueba',
                'fecha_insercion': datetime.now()
            }
        ]
        
        test_stats = {
            'total_products': 2,
            'avg_price': 24.99,
            'category_stats': [
                {'category': 'electronics', 'count': 1, 'avg_price': 19.99, 'min_price': 19.99, 'max_price': 19.99},
                {'category': 'clothing', 'count': 1, 'avg_price': 29.99, 'min_price': 29.99, 'max_price': 29.99}
            ]
        }
        
        # Generar reporte
        filepath = generator.generate_report(test_products, test_stats)
        logger.info(f"✅ Test exitoso: {filepath}")
        
        return True
        
    except Exception as e:
        logger.error(f"❌ Test fallido: {str(e)}")
        return False


if __name__ == "__main__":
    test_excel_generator()
